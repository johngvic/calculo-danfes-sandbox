#!/usr/bin/env python3
import argparse
import csv
from openpyxl import load_workbook

def main():
  p = argparse.ArgumentParser(description="Convert XLSX to CSV (streaming)")
  p.add_argument("input")
  p.add_argument("output")
  p.add_argument("--sheet", default="0", help="sheet name or index (default 0)")
  p.add_argument("--sep", default="¦", help="CSV delimiter")
  p.add_argument("--encoding", default="utf-8", help="output encoding")
  args = p.parse_args()

  wb = load_workbook(filename=args.input, data_only=True)
  sheet = int(args.sheet) if str(args.sheet).isdigit() else args.sheet
  ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]

  with open(args.output, "w", newline="", encoding=args.encoding) as f:
    writer = csv.writer(f, delimiter=args.sep)
    for row in ws.iter_rows(values_only=True):
      writer.writerow([("" if v is None else v) for v in row])

if __name__ == "__main__":
  main()